import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import io
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series, BarChart

# 1. 設定網頁標題與配置
st.set_page_config(page_title="氣密測試數據自動化分析看板", layout="wide")
st.title("📊 氣密測試 (Leak Test) 數據自動化看板")

# 2. 頂部主選單
app_mode = st.radio(
    "請選擇分析模式：",
    ["📈 詳細趨勢曲線 (Line Plot Mode)", "📊 整體數據分佈 (Box Plot Mode)"],
    horizontal=True
)

st.write("---")

# ==============================================================================
# 模式一：Line Plot 趨勢曲線 (第一段：資料流讀取與前置矩陣運算)
# ==============================================================================
if app_mode == "📈 詳細趨勢曲線 (Line Plot Mode)":
    st.subheader("趨勢曲線分析 (Line Plot)")
    
    st.sidebar.header("⚙️ 趨勢圖控制")
    sample_step = st.sidebar.slider("📊 數據抽樣間隔 (Step)", min_value=1, max_value=10, value=1)
    
    uploaded_file = st.file_uploader("請上傳詳細製程數據 CSV 檔案 (Line Plot)", type=["csv"], key="line_uploader")
    
    if uploaded_file is not None:
        try:
            df = pd.read_csv(uploaded_file)
            required_columns = ['SN', 'Timestamp', 'Pressure(Kpa)', 'Leak']
            missing_cols = [col for col in required_columns if col not in df.columns]
            
            if missing_cols:
                st.error(f"❌ 檔案格式不符，缺少必要欄位: {missing_cols}")
            else:
                df['Timestamp'] = pd.to_datetime(df['Timestamp'])
                
                # 自動忽略 Phase 包含 'Stabilization'
                if 'Phase' in df.columns:
                    initial_count = len(df)
                    df = df[~df['Phase'].astype(str).str.contains('Stabilization', case=False, na=False)]
                    filtered_count = initial_count - len(df)
                    st.sidebar.warning(f"🧹 已自動過濾掉 {filtered_count} 筆 Stabilization 數據。")
                
                df = df.sort_values(by='Timestamp')
                unique_sns = df['SN'].dropna().unique()
                st.success(f"✅ 成功讀取數據！偵測到共 {len(unique_sns)} 個不同的 SN 產品。")
                
                # 建構多分頁 Excel 檔案
                output_excel = io.BytesIO()
                wb = Workbook()
                wb.remove(wb.active)  # 移除預設工作表
                for sn in unique_sns:
                    sn_data = df[df['SN'] == sn]
                    short_sn = str(sn).split(':')[-1] if ':' in str(sn) else str(sn)
                    sheet_name = short_sn[-30:]
                    
                    ws = wb.create_sheet(title=sheet_name)
                    headers = ['Index', 'Timestamp', 'Pressure(Kpa)', 'Leak', 'bResult']
                    ws.append(headers)
                    
                    for idx, (_, row) in enumerate(sn_data.iterrows(), start=1):
                        ws.append([
                            idx,
                            str(row['Timestamp']), 
                            row['Pressure(Kpa)'], 
                            row['Leak'] if pd.notna(row['Leak']) else "", 
                            row['bResult'] if 'bResult' in df.columns else ""
                        ])
                    
                    num_rows = len(sn_data)
                    
                    # 計算動態最大值與最小值緩衝，留出刻度文字空間
                    raw_max_p = float(sn_data['Pressure(Kpa)'].max()) if not sn_data['Pressure(Kpa)'].dropna().empty else 100.0
                    raw_min_p = float(sn_data['Pressure(Kpa)'].min()) if not sn_data['Pressure(Kpa)'].dropna().empty else 0.0
                    raw_max_l = float(sn_data['Leak'].max()) if not sn_data['Leak'].dropna().empty else 1.0
                    raw_min_l = float(sn_data['Leak'].min()) if not sn_data['Leak'].dropna().empty else 0.0
                    
                    max_p = raw_max_p + abs(raw_max_p * 0.05) if raw_max_p != 0 else 10.0
                    min_p = raw_min_p - abs(raw_min_p * 0.05) if raw_min_p != 0 else -10.0
                    max_l = raw_max_l + abs(raw_max_l * 0.10) if raw_max_l != 0 else 1.0
                    min_l = raw_min_l - abs(raw_min_l * 0.10) if raw_min_l != 0 else -0.1
                    
                    # ----------------- 建立 Excel 壓力高質感深色散佈圖 -----------------
                    chart_p = ScatterChart()
                    chart_p.title = f"SN {sheet_name} - Pressure Trend"
                    
                    # 直接指定 Excel 的內建專業黑底科技風樣式 (Style 32)
                    chart_p.style = 32 
                    chart_p.width = 26   
                    chart_p.height = 14  
                    
                    chart_p.x_axis.title = "Data Point Index"
                    chart_p.y_axis.title = "Pressure (Kpa)"
                    
                    # 鎖定上下限範圍，給數值刻度最精準的伸展空間
                    chart_p.y_axis.scaling.max = max_p
                    chart_p.y_axis.scaling.min = min_p
                    
                    # 強制開啟主網格線與坐標軸可見度（確保文字刻度 100% 顯現）
                    chart_p.y_axis.majorGridlines = True
                    chart_p.x_axis.majorGridlines = True
                    chart_p.x_axis.tickLblPos = "nextTo"
                    chart_p.y_axis.tickLblPos = "nextTo"
                    
                    chart_p.legend = None  
                    
                    x_values_p = Reference(ws, min_col=1, min_row=2, max_row=num_rows+1)
                    y_values_p = Reference(ws, min_col=3, min_row=1, max_row=num_rows+1)
                    
                    series_p = Series(y_values_p, x_values_p, title_from_data=True)
                    series_p.graphicalProperties.line.solidFill = "00FF00" # 配上亮眼綠色曲線
                    series_p.graphicalProperties.line.width = 25000       
                    
                    chart_p.append(series_p)
                    ws.add_chart(chart_p, "G2")
                    
                    # ----------------- 建立 Excel 洩漏率高質感深色散佈圖 -----------------
                    if not sn_data['Leak'].dropna().empty:
                        chart_l = ScatterChart()
                        chart_l.title = f"SN {sheet_name} - Leak Trend"
                        
                        # 同步鎖定黑底科技風樣式
                        chart_l.style = 32
                        chart_l.width = 26
                        chart_l.height = 14
                        
                        chart_l.x_axis.title = "Data Point Index"
                        chart_l.y_axis.title = "Leak Value"
                        
                        chart_l.y_axis.scaling.max = max_l
                        chart_l.y_axis.scaling.min = min_l
                        
                        chart_l.y_axis.majorGridlines = True
                        chart_l.x_axis.majorGridlines = True
                        chart_l.x_axis.tickLblPos = "nextTo"
                        chart_l.y_axis.tickLblPos = "nextTo"
                        
                        chart_l.legend = None
                        
                        x_values_l = Reference(ws, min_col=1, min_row=2, max_row=num_rows+1)
                        y_values_l = Reference(ws, min_col=4, min_row=1, max_row=num_rows+1)
                        
                        series_l = Series(y_values_l, x_values_l, title_from_data=True)
                        series_l.graphicalProperties.line.solidFill = "00FF00" 
                        series_l.graphicalProperties.line.width = 25000
                        
                        chart_l.append(series_l)
                        ws.add_chart(chart_l, "G18") 
                wb.save(output_excel)
                output_excel.seek(0)
                
                st.download_button(
                    label="📥 一鍵下載自動生成的『多分頁含圖表』Excel 報告",
                    data=output_excel,
                    file_name="Leak_Test_Line_Plots.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.write("---")
                
                # 前端 Tabs 網頁渲染
                tabs = st.tabs([f"SN: {sn.split(':')[-1] if ':' in str(sn) else sn}" for sn in unique_sns])
                for i, sn in enumerate(unique_sns):
                    with tabs[i]:
                        st.subheader(f"產品序號: {sn}")
                        sn_data = df[df['SN'] == sn]
                        if sn_data.empty:
                            st.info("ℹ️ 該產品在排除 Stabilization 階段後無剩餘數據。")
                            continue
                        
                        plot_data = sn_data.iloc[::sample_step] if sample_step > 1 else sn_data
                        
                        col1, col2, col3 = st.columns(3)
                        col1.metric("有效數據筆數", f"{len(sn_data)} 筆")
                        col2.metric("最大壓力 (Kpa)", f"{sn_data['Pressure(Kpa)'].max():.2f}")
                        col3.metric("測試結果狀態", f"{sn_data['bResult'].iloc[-1] if 'bResult' in sn_data.columns else '未知'}")
                        
                        st.markdown("#### 📈 Pressure (Kpa) 趨勢變化圖")
                        fig_press = go.Figure()
                        fig_press.add_trace(go.Scatter(x=plot_data['Timestamp'], y=plot_data['Pressure(Kpa)'], name='Pressure', mode='lines+markers', line=dict(color='#1f77b4', width=2), marker=dict(size=5)))
                        fig_press.update_layout(xaxis=dict(title="Timestamp"), yaxis=dict(title="Pressure (Kpa)"), hovermode="closest")
                        st.plotly_chart(fig_press, use_container_width=True)
                        
                        st.markdown("#### 📉 Leak 趨勢變化圖")
                        if not sn_data['Leak'].dropna().empty:
                            fig_leak = go.Figure()
                            fig_leak.add_trace(go.Scatter(x=plot_data['Timestamp'], y=plot_data['Leak'], name='Leak', mode='lines+markers', line=dict(color='#ff7f0e', width=2), marker=dict(size=5)))
                            fig_leak.update_layout(xaxis=dict(title="Timestamp"), yaxis=dict(title="Leak"), hovermode="closest")
                            st.plotly_chart(fig_leak, use_container_width=True)
                        else:
                            st.info("ℹ️ 此 SN 當前數據中的 Leak 欄位皆為空值。")
                            
        except Exception as e:
            st.error(f"讀取檔案時發生錯誤: {e}")

# ==============================================================================
# 模式二：Box Plot 箱線圖
# ==============================================================================
elif app_mode == "📊 整體數據分佈 (Box Plot Mode)":
    st.subheader("整體數據箱線圖分析 (Box Plot)")
    
    uploaded_file_box = st.file_uploader("請上傳統計數據 CSV 檔案 (Box Plot)", type=["csv"], key="box_uploader")
    
    if uploaded_file_box is not None:
        try:
            df_box = pd.read_csv(uploaded_file_box)
            required_box_cols = ['SN', 'Pressure(Kpa)', 'Leak']
            missing_box_cols = [col for col in required_box_cols if col not in df_box.columns]
            
            if missing_box_cols:
                st.error(f"❌ 檔案格式不符，缺少必要欄位: {missing_box_cols}")
            else:
                st.success(f"✅ 成功讀取統計數據！共有 {len(df_box)} 筆產品總點數。")
                
                if 'bResult' in df_box.columns:
                    ng_count = len(df_box[df_box['bResult'] == 'NG'])
                    ok_count = len(df_box[df_box['bResult'] == 'OK'])
                    col_b1, col_b2 = st.columns(2)
                    col_b1.metric("🟢 OK 總數", f"{ok_count} 筆")
                    col_b2.metric("🔴 NG 總數", f"{ng_count} 筆")
                
                output_box_excel = io.BytesIO()
                wb_box = Workbook()
                
                ws_data = wb_box.active
                ws_data.title = "Raw_Data"
                ws_data.append(list(df_box.columns))
                for _, row in df_box.iterrows():
                    ws_data.append(list(row))
                
                ws_chart = wb_box.create_sheet(title="Excel_Charts")
                ws_chart.append(["💡 提示：此工作表右側已為您安插高相容性的 Pressure 與 Leak 數值對比圖表。"])
                
                p_col_idx = df_box.columns.get_loc('Pressure(Kpa)') + 1
                l_col_idx = df_box.columns.get_loc('Leak') + 1
                
                chart_box_p = BarChart()
                chart_box_p.type = "col"
                chart_box_p.style = 10
                chart_box_p.title = "Pressure (Kpa) Summary"
                chart_box_p.y_axis.title = "Pressure (Kpa)"
                data_box_p = Reference(ws_data, min_col=p_col_idx, min_row=1, max_row=len(df_box)+1)
                chart_box_p.add_data(data_box_p, titles_from_data=True)
                chart_box_p.legend = None
                ws_chart.add_chart(chart_box_p, "C3")
                
                chart_box_l = BarChart()
                chart_box_l.type = "col"
                chart_box_l.style = 11
                chart_box_l.title = "Leak Value Summary"
                chart_box_l.y_axis.title = "Leak Value"
                data_box_l = Reference(ws_data, min_col=l_col_idx, min_row=1, max_row=len(df_box)+1)
                chart_box_l.add_data(data_box_l, titles_from_data=True)
                chart_box_l.legend = None
                ws_chart.add_chart(chart_box_l, "K3")
                
                wb_box.save(output_box_excel)
                output_box_excel.seek(0)
                
                st.download_button(
                    label="📥 點此下載多產品品質數據 Excel 報告",
                    data=output_box_excel,
                    file_name="Leak_Test_Quality_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.write("---")
                
                graph_col1, graph_col2 = st.columns(2)
                with graph_col1:
                    st.markdown("#### 📦 Pressure (Kpa) 箱線分佈圖")
                    fig_box_press = px.box(df_box, y="Pressure(Kpa)", points="all", hover_data=["SN"], color="bResult" if "bResult" in df_box.columns else None, color_discrete_map={"OK": "#1f77b4", "NG": "#ef553b"})
                    st.plotly_chart(fig_box_press, use_container_width=True)
                    
                with graph_col2:
                    st.markdown("#### 📦 Leak 箱線分佈圖")
                    fig_box_leak = px.box(df_box, y="Leak", points="all", hover_data=["SN"], color="bResult" if "bResult" in df_box.columns else None, color_discrete_map={"OK": "#1f77b4", "NG": "#ef553b"})
                    st.plotly_chart(fig_box_leak, use_container_width=True)
                    
        except Exception as e:
            st.error(f"讀取 Box Plot 檔案時發生錯誤: {e}")
